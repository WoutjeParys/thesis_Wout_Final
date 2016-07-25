__author__ = 'Wokkie'

import os

from pandas import pivot_table, merge, ExcelWriter, DataFrame
import numpy as np

#import some usefull things
from openpyxl import Workbook
from openpyxl import load_workbook

from gams_addon import gdx_to_df, DomainInfo

from openpyxl.styles import Style, Border, Alignment, Protection, Font, colors


file = 'results\out_db_40_DR.gdx'
gdx_file = os.path.join(os.getcwd(), '%s' % file)
writefile = os.getcwd() + '\\' + 'excel\output_elasticity_model.xlsx'
writer = ExcelWriter(writefile)

print gdx_file
zone_dict = dict()
zone_dict['BEL_Z'] = 'BEL'

print 'Retrieving price_unit_clone'
price_unit = gdx_to_df(gdx_file, 'price_unit_clone')
old_index = price_unit.index.names
price_unit['C'] = [zone_dict[z] for z in price_unit.index.get_level_values('Z')]
price_unit.set_index('C', append=True, inplace=True)
price_unit = price_unit.reorder_levels(['C'] + old_index)
price_unit.reset_index(inplace=True)
price_unit = pivot_table(price_unit, 'price_unit_clone', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)

print 'Retrieving demand_unit'
demand_unit = gdx_to_df(gdx_file, 'demand_unit')
old_index = demand_unit.index.names
demand_unit['C'] = [zone_dict[z] for z in demand_unit.index.get_level_values('Z')]
demand_unit.set_index('C', append=True, inplace=True)
demand_unit = demand_unit.reorder_levels(['C'] + old_index)
demand_unit.reset_index(inplace=True)
demand_unit = pivot_table(demand_unit, 'demand_unit', index=['P','T','Z'], columns=['C'], aggfunc=np.sum)
# print curt.head()

print 'Retrieving demand_ref'
demand_ref = gdx_to_df(gdx_file, 'demand_ref')
old_index = demand_ref.index.names
demand_ref['C'] = [zone_dict[z] for z in demand_ref.index.get_level_values('Z')]
demand_ref.set_index('C', append=True, inplace=True)
demand_ref = demand_ref.reorder_levels(['C'] + old_index)
demand_ref.reset_index(inplace=True)
demand_ref = pivot_table(demand_ref, 'demand_ref', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)

print 'Retrieving demand_new_res'
demand_new_res = gdx_to_df(gdx_file, 'demand_new_res')
old_index = demand_new_res.index.names
demand_new_res['C'] = [zone_dict[z] for z in demand_new_res.index.get_level_values('Z')]
demand_new_res.set_index('C', append=True, inplace=True)
demand_new_res = demand_new_res.reorder_levels(['C'] + old_index)
demand_new_res.reset_index(inplace=True)
demand_new_res = pivot_table(demand_new_res, 'demand_new_res', index=['P','T','Z'], columns=['C'], aggfunc=np.sum)
# print curt.head()

print 'Retrieving demand_res_ref'
# DEM_REF_RES = gdx_to_df(gdx_file, 'DEM_REF_RES')
DEM_REF_RES = gdx_to_df(gdx_file, 'DEM_OPTIMAL')
old_index = DEM_REF_RES.index.names
DEM_REF_RES['C'] = [zone_dict[z] for z in DEM_REF_RES.index.get_level_values('Z')]
DEM_REF_RES.set_index('C', append=True, inplace=True)
DEM_REF_RES = DEM_REF_RES.reorder_levels(['C'] + old_index)
DEM_REF_RES.reset_index(inplace=True)
# DEM_REF_RES = pivot_table(DEM_REF_RES, 'DEM_REF_RES', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)
DEM_REF_RES = pivot_table(DEM_REF_RES, 'DEM_OPTIMAL', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)

print 'Retrieving demand_res_min'
DEM_RES_MIN = gdx_to_df(gdx_file, 'DEM_RES_MIN')
old_index = DEM_RES_MIN.index.names
DEM_RES_MIN['C'] = [zone_dict[z] for z in DEM_RES_MIN.index.get_level_values('Z')]
DEM_RES_MIN.set_index('C', append=True, inplace=True)
DEM_RES_MIN = DEM_RES_MIN.reorder_levels(['C'] + old_index)
DEM_RES_MIN.reset_index(inplace=True)
# DEM_REF_RES = pivot_table(DEM_REF_RES, 'DEM_REF_RES', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)
DEM_RES_MIN = pivot_table(DEM_RES_MIN, 'DEM_RES_MIN', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)

print 'Retrieving demand_res_max'
DEM_RES_MAX = gdx_to_df(gdx_file, 'DEM_RES_MAX')
old_index = DEM_RES_MAX.index.names
DEM_RES_MAX['C'] = [zone_dict[z] for z in DEM_RES_MAX.index.get_level_values('Z')]
DEM_RES_MAX.set_index('C', append=True, inplace=True)
DEM_RES_MAX = DEM_RES_MAX.reorder_levels(['C'] + old_index)
DEM_RES_MAX.reset_index(inplace=True)
# DEM_REF_RES = pivot_table(DEM_REF_RES, 'DEM_REF_RES', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)
DEM_RES_MAX = pivot_table(DEM_RES_MAX, 'DEM_RES_MAX', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)

print 'Retrieving demand_res_max'
DEM_RES_FP = gdx_to_df(gdx_file, 'DEM_RES_FP')
old_index = DEM_RES_FP.index.names
DEM_RES_FP['C'] = [zone_dict[z] for z in DEM_RES_FP.index.get_level_values('Z')]
DEM_RES_FP.set_index('C', append=True, inplace=True)
DEM_RES_FP = DEM_RES_FP.reorder_levels(['C'] + old_index)
DEM_RES_FP.reset_index(inplace=True)
# DEM_REF_RES = pivot_table(DEM_REF_RES, 'DEM_REF_RES', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)
DEM_RES_FP = pivot_table(DEM_RES_FP, 'DEM_RES_FP', index=['P', 'T','Z'], columns=['C'], aggfunc=np.sum)


# First Merge
genmarg = merge(demand_unit, demand_ref, left_index=True, right_index=True, how='outer', suffixes=['_dem', '_dem_ref'])
genmargres = merge(demand_new_res,DEM_REF_RES,left_index=True,right_index=True,how='outer',suffixes=['dem_res','_dem_res_ref'])
genmarg = merge(genmarg, genmargres, left_index=True, right_index=True, how='outer')
genmarg = merge(genmarg, DEM_RES_MIN, left_index=True, right_index=True, how='outer')
genmarg = merge(genmarg, DEM_RES_MAX, left_index=True, right_index=True, how='outer')
genmarg = merge(genmarg, DEM_RES_FP, left_index=True, right_index=True, how='outer')
genmarg = merge(genmarg, price_unit, left_index=True, right_index=True, how='outer', suffixes=['', '_price'])

print 'Writing demand and prices to Excel'
genmarg.to_excel(writer, na_rep=0.0, sheet_name='pattern', merge_cells=False)

# print 'Retrieving innerframe'
# innerframe = gdx_to_df(gdx_file, 'innerframe')
# old_index = innerframe.index.names
# innerframe['C'] = [zone_dict[z] for z in innerframe.index.get_level_values('Z')]
# innerframe.set_index('C', append=True, inplace=True)
# innerframe = innerframe.reorder_levels(['C'] + old_index)
# innerframe.reset_index(inplace=True)
# innerframe = pivot_table(innerframe, 'innerframe', index=['P', 'H','Z'], columns=['C'], aggfunc=np.sum)
#
# print 'Retrieving outerframe'
# outerframe = gdx_to_df(gdx_file, 'outerframe')
# old_index = outerframe.index.names
# outerframe['C'] = [zone_dict[z] for z in outerframe.index.get_level_values('Z')]
# outerframe.set_index('C', append=True, inplace=True)
# outerframe = outerframe.reorder_levels(['C'] + old_index)
# outerframe.reset_index(inplace=True)
# outerframe = pivot_table(outerframe, 'outerframe', index=['P', 'H','Z'], columns=['C'], aggfunc=np.sum)
#
# framemerg = merge(innerframe,outerframe, left_index=True, right_index=True, how='outer', suffixes=['_inner', '_outer'])
# framemerg.to_excel(writer, na_rep=0.0, sheet_name='frames', merge_cells=False)

print 'Creating generation pattern analysis'
print 'Retrieving gen'
gen = gdx_to_df(gdx_file, 'gen')
old_index = gen.index.names
gen['C'] = [zone_dict[z] for z in gen.index.get_level_values('Z')]
gen.set_index('C', append=True, inplace=True)
gen = gen.reorder_levels(['C'] + old_index)
gen.reset_index(inplace=True)
gen = pivot_table(gen, 'gen', index=['C', 'Y', 'P', 'T'], columns=['G'], aggfunc=np.sum)

print 'Writing pattern to Excel'
gen.to_excel(writer, na_rep=0.0, sheet_name='gen_pattern', merge_cells=False)


print 'get capacities and objective'
print 'retrieving cap'
cap = gdx_to_df(gdx_file, 'cap')
# print cap
old_index = cap.index.names
cap['C'] = [zone_dict[z] for z in cap.index.get_level_values('Z')]
cap.set_index('C', append=True, inplace=True)
cap = cap.reorder_levels(['C'] + old_index)
cap.reset_index(inplace=True)
cap = pivot_table(cap, 'cap', index=['Y', 'Z', 'G'], columns=['C'], aggfunc=np.sum)
cap.to_excel(writer, na_rep=0.0, sheet_name='capacities', merge_cells=False)

print 'retrieving cost'
cost = gdx_to_df(gdx_file, 'obj')
# print cost
old_index = cost.index.names
# cost['C'] = [zone_dict[z] for z in cost.index.get_level_values('Z')]
# cost.set_index('C', append=True, inplace=True)
# cost = cost.reorder_levels(['C'] + old_index)
# cost.reset_index(inplace=True)
# cost = pivot_table(cost, 'totalcost', index=[], columns=['C'], aggfunc=np.sum)
# print cost

print 'Writing objective to Excel'
cost.to_excel(writer, na_rep=0.0, sheet_name='objective', merge_cells=False)

print 'get balance'
print 'retrieving marg'
marg = gdx_to_df(gdx_file, 'marg')
old_index = marg.index.names
marg['C'] = [zone_dict[z] for z in marg.index.get_level_values('Z')]
marg.set_index('C', append=True, inplace=True)
marg = marg.reorder_levels(['C'] + old_index)
marg.reset_index(inplace=True)
marg = pivot_table(marg, 'marg', index=['Y', 'P', 'T'], columns=['C'], aggfunc=np.sum)

print 'Writing balances.m to Excel'
marg.to_excel(writer, na_rep=0.0, sheet_name='balance', merge_cells=False)


writer.close()

# wb = load_workbook(writefile)
# ws1 = wb.active
# gen_techn = list()
# gen_energ = list()
# gen_margc = list()
# final = list()
# for r in range (2,len(ws1.rows)+1,1):
# #smaller loop for testing
# #for r in range (2,100,1):
#     currentg = ws1.cell(row = r, column = 4).value
#     currente = ws1.cell(row = r, column = 5).value
#     currentc = ws1.cell(row = r, column = 6).value
#     if currentg not in gen_techn:
#         gen_techn.append(currentg)
#         gen_energ.append(currente)
#         gen_margc.append(currentc)
#     else:
#         max_cost = 0
#         amount_techn = len(gen_techn)
#         for i in range(0,amount_techn-1,1):
#             #TODO
#             #if gen_energ[i] != '0.0'  and gen_margc[i] > max_cost:
#             if gen_energ[i] != 0 and gen_energ[i] != '0.0' and gen_margc[i] > max_cost:
#                 max_cost = gen_margc[i]
#         final.append([ws1.cell(row = r-1, column = 1).value,ws1.cell(row = r-1, column = 2).value,ws1.cell(row = r-1, column = 3).value,max_cost])
#         gen_techn = list()
#         gen_energ = list()
#         gen_margc = list()
#         gen_techn.append(currentg)
#         gen_energ.append(currente)
#         gen_margc.append(currentc)
#     if r == len(ws1.rows):
#         max_cost = 0
#         amount_techn = len(gen_techn)
#         for i in range(0,amount_techn-1,1):
#             #if gen_energ[i] != '0.0'  and gen_margc[i] > max_cost:
#             if gen_energ[i] != 0 and gen_energ[i] != '0.0' and gen_margc[i] > max_cost:
#                 max_cost = gen_margc[i]
#         final.append([ws1.cell(row = r-1, column = 1).value,ws1.cell(row = r-1, column = 2).value,ws1.cell(row = r-1, column = 3).value,max_cost])
#         gen_techn = list()
#         gen_energ = list()
#         gen_margc = list()
#         gen_techn.append(currentg)
#         gen_energ.append(currente)
#         gen_margc.append(currentc)
#
# headings = Style(font=Font(size=12,bold=True,color=colors.RED))
# ws2 = wb.create_sheet(title='prices')
# total_rows = len(final)
# for i in range(1,total_rows+1,1):
#     for j in range (1,5,1):
#         if i == 1:
#             c1 = ws2.cell(row=i, column=j)
#             if j == 4:
#                 c1.value = 'marg_price'
#             else:
#                 c1.value = ws1.cell(row=i,column=j).value
#             c1.style = headings
#         c = ws2.cell(row=i+1, column=j)
#         c.value = final[i-1][j-1]
# wb.save(writefile)






