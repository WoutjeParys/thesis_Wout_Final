__author__ = 'Wout'

import sqlite3 as sq
import csv
import xlrd
import os
from matplotlib import pyplot as plt
import openpyxl


include = False
include_when_period_changed = True

geographical_entities = include
generation_technologies = include
storage_technologies = include
time_periods = include_when_period_changed
demand_ref = include
intermittent_renewables = include
reliable_intermittent = include
policy = include
installed_capacities = include
reserves_techn = include
fixed_tariffs = include
elasticity_matrix = include_when_period_changed
shiftingconstraints = include_when_period_changed
correction_matrix = include_when_period_changed
price_profiles = include_when_period_changed
demand_profiles = include_when_period_changed

def initialise(length_period):
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()

    print os.getcwd()


    if geographical_entities:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Geographical_entities.xlsx"))
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS Region;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Region (Code TEXT, description TEXT);'
        # ,  PRIMARY KEY(Code));'
        cur.execute(sql)
        regions = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            # print a, b
            regions.append((a, b))
        # print areas
        cur.executemany('INSERT INTO Region VALUES (?,?)', regions)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(1)
        sql = 'DROP TABLE IF EXISTS Country;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Country (Include BOOLEAN, Code TEXT, Region_Code TEXT, Member_State TEXT, description TEXT);'
        # ,  PRIMARY KEY(Code));'
        cur.execute(sql)
        countries = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            c = sh.cell_value(rowx=rx + 3, colx=2)
            d = sh.cell_value(rowx=rx + 3, colx=3)
            e = sh.cell_value(rowx=rx + 3, colx=4)
            # print a,b
            countries.append((a, b, c, d, e))
        # print areas
        cur.executemany('INSERT INTO Country VALUES (?,?,?,?,?)', countries)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(2)
        sql = 'DROP TABLE IF EXISTS Zone;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Zone (Code TEXT, Country_Code TEXT, description TEXT);'
        # ,  PRIMARY KEY(Code));'
        cur.execute(sql)
        zones = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            c = sh.cell_value(rowx=rx + 3, colx=2)
            # print a,b
            zones.append((a, b, c))
        # print areas
        cur.executemany('INSERT INTO Zone VALUES (?,?,?)', zones)
        conn.commit()
        ############################################
        print "Done geographical_entities"

    if generation_technologies:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Generation_technologies.xlsx"))
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS Generation_technologies;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Generation_technologies (Include BOOLEAN, Dispatchable BOOLEAN, Gas BOOLEAN, RES BOOLEAN, Code TEXT, description TEXT);'
        cur.execute(sql)
        generation = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            c = sh.cell_value(rowx=rx + 3, colx=2)
            d = sh.cell_value(rowx=rx + 3, colx=3)
            e = sh.cell_value(rowx=rx + 3, colx=4)
            f = sh.cell_value(rowx=rx + 3, colx=5)
            # print a, b, c
            generation.append((a, b, c, d, e, f))
        # print areas
        cur.executemany('INSERT INTO Generation_technologies VALUES (?,?,?,?,?,?)', generation)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(1)
        sql = 'DROP TABLE IF EXISTS Generation_parameters;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Generation_parameters (Code TEXT, description TEXT);'
        cur.execute(sql)
        parameters = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            # print a, b
            parameters.append((a, b))
        #print parameters
        cur.executemany('INSERT INTO Generation_parameters VALUES (?,?)', parameters)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(2)
        sql = 'DROP TABLE IF EXISTS Generation_data;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Generation_data (Technology TEXT, Parameter TEXT, Value FLOAT);'
        cur.execute(sql)
        gdata = list()
        technologies = [tech[0] for tech in cur.execute('SELECT Code from Generation_technologies').fetchall()]
        # print technologies
        parameters = [param[0] for param in cur.execute('SELECT Code from Generation_parameters').fetchall()]
        # print parameters
        for row in range(3, sh.nrows):
            tech = sh.cell_value(row, 0)
            if tech in technologies:
                for col in range(1, sh.ncols):
                    param = sh.cell_value(2, col)
                    if param in parameters:
                        value = sh.cell_value(row, col)
                        # print tech, param, value
                        gdata.append((tech, param, value))
        cur.executemany('INSERT INTO Generation_data VALUES (?,?,?)', gdata)
        conn.commit()
        ############################################
        print "Done generation_technologies"

    if storage_technologies:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Storage_technologies.xlsx"))
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS Storage_technologies;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Storage_technologies (Include BOOLEAN, Short BOOLEAN, Mid BOOLEAN, Long BOOLEAN, Code TEXT, description TEXT);'
        cur.execute(sql)
        storage = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            c = sh.cell_value(rowx=rx + 3, colx=2)
            d = sh.cell_value(rowx=rx + 3, colx=3)
            e = sh.cell_value(rowx=rx + 3, colx=4)
            f = sh.cell_value(rowx=rx + 3, colx=5)
            # print a, b, c
            storage.append((a, b, c, d, e, f))
        # print areas
        cur.executemany('INSERT INTO Storage_technologies VALUES (?,?,?,?,?,?)', storage)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(1)
        sql = 'DROP TABLE IF EXISTS Storage_parameters;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Storage_parameters (Code TEXT, description TEXT);'
        cur.execute(sql)
        parameters = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            # print a, b
            parameters.append((a, b))
        #print parameters
        cur.executemany('INSERT INTO Storage_parameters VALUES (?,?)', parameters)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(2)
        sql = 'DROP TABLE IF EXISTS Storage_data;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Storage_data (Technology TEXT, Parameter TEXT, Value FLOAT);'
        cur.execute(sql)
        sdata = list()
        technologies = [tech[0] for tech in cur.execute('SELECT Code from Storage_technologies').fetchall()]
        # print technologies
        parameters = [param[0] for param in cur.execute('SELECT Code from Storage_parameters').fetchall()]
        # print parameters
        for row in range(3, sh.nrows):
            tech = sh.cell_value(row, 0)
            if tech in technologies:
                for col in range(1, sh.ncols):
                    param = sh.cell_value(2, col)
                    if param in parameters:
                        value = sh.cell_value(row, col)
                        # print tech, param, value
                        sdata.append((tech, param, value))
        cur.executemany('INSERT INTO Storage_data VALUES (?,?,?)', sdata)
        conn.commit()
        ############################################
        print "Done storage_technologies"

    if time_periods:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Time_periods.xlsx"))
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS Years;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Years (Include BOOLEAN, Year TEXT);'
        # ,  PRIMARY KEY(Code));'
        cur.execute(sql)
        year = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            # print a, int(b)
            year.append((a, int(b)))
        # print year
        cur.executemany('INSERT INTO Years VALUES (?,?)', year)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(1)
        sql = 'DROP TABLE IF EXISTS Time_steps;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Time_steps (First_hour FLOAT, Weight FLOAT, Season FLOAT);'
        cur.execute(sql)
        periods = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            s = sh.cell_value(rowx=rx + 3, colx=2)
            # print a, b
            periods.append((a, b, s))
        # print periods
        cur.executemany('INSERT INTO Time_steps VALUES (?,?,?)', periods)
        conn.commit()
        ############################################
        print "Done time"

    startday_weekend = 2

    if price_profiles:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\DemR\PriceProf.xlsx"))
        sql = 'DROP TABLE IF EXISTS PriceProfile;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS PriceProfile (Season FLOAT, Zone TEXT, Hour TEXT, Price FLOAT);'
        cur.execute(sql)
        prices = list()
        zone = 'BEL_Z'
        amount_of_days = length_period/24
        for season in range (0,4):
            print 'season: ', season
            for day in range(0,amount_of_days):
                if day == startday_weekend or day == startday_weekend+1:
                    print 'weekendday with sheet index = ', season*2+1
                    sh=book.sheet_by_index(season*2+1)
                else:
                    print 'weekday with sheet index = ', season*2
                    sh=book.sheet_by_index(season*2)
                for col in range(1,sh.ncols):
                    hour = int(sh.cell_value(0,col)) + 24*day
                    value = sh.cell_value(1,col)
                    prices.append((season+1,zone,hour,value))
        cur.executemany('INSERT INTO PriceProfile VALUES (?,?,?,?)', prices)
        conn.commit()
        print 'Done price profiles'

    if demand_profiles:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\DemR\DemAllProfilesFull.xlsx"))
        sqlref = 'DROP TABLE IF EXISTS Dem_ref_profile;'
        sqlmin = 'DROP TABLE IF EXISTS Dem_min_profile;'
        sqlmax = 'DROP TABLE IF EXISTS Dem_max_profile;'
        sqlflat = 'DROP TABLE IF EXISTS Dem_flat_profile;'
        cur.execute(sqlref)
        cur.execute(sqlmin)
        cur.execute(sqlflat)
        cur.execute(sqlmax)
        sqlref = 'CREATE TABLE IF NOT EXISTS Dem_ref_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
        sqlmin = 'CREATE TABLE IF NOT EXISTS Dem_min_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
        sqlmax = 'CREATE TABLE IF NOT EXISTS Dem_max_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
        sqlflat = 'CREATE TABLE IF NOT EXISTS Dem_flat_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
        cur.execute(sqlref)
        cur.execute(sqlmin)
        cur.execute(sqlmax)
        cur.execute(sqlflat)
        demref = list()
        demmin = list()
        demmax = list()
        demflp = list()
        zone = 'BEL_Z'
        shmin=book.sheet_by_index(0)
        shmax=book.sheet_by_index(1)
        shref=book.sheet_by_index(2)
        shflp=book.sheet_by_index(3)
        amount_of_days = length_period/24
        for season in range (0,4):
            print 'season: ', season
            for day in range(0,amount_of_days):
                if day == startday_weekend or day == startday_weekend+1:
                    print 'weekendday with row = ', season*2+1
                    row = season*2+2
                else:
                    print 'weekday with sheet index = ', season*2
                    row = season*2+1
                for col in range(1,shref.ncols):
                    hour = int(shref.cell_value(0,col)) + 24*day
                    valueref = shref.cell_value(row,col)
                    valuemin = shmin.cell_value(row,col)
                    valuemax = shmax.cell_value(row,col)
                    valueflp = shflp.cell_value(row,col)
                    demref.append((season+1,zone,hour,valueref))
                    demmin.append((season+1,zone,hour,valuemin))
                    demmax.append((season+1,zone,hour,valuemax))
                    demflp.append((season+1,zone,hour,valueflp))
        cur.executemany('INSERT INTO Dem_ref_profile VALUES (?,?,?,?)', demref)
        cur.executemany('INSERT INTO Dem_min_profile VALUES (?,?,?,?)', demmin)
        cur.executemany('INSERT INTO Dem_max_profile VALUES (?,?,?,?)', demmax)
        cur.executemany('INSERT INTO Dem_flat_profile VALUES (?,?,?,?)', demflp)
        conn.commit()
        print 'Done price profiles'

    #test generated elasticity & diag/triag matrices
    test = False
    if test:
        wbwrite = openpyxl.load_workbook('excel/TestElasticity.xlsx')

    if elasticity_matrix:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Elasticity.xlsx"))
        sql = 'DROP TABLE IF EXISTS Elasticity;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Elasticity (Season FLOAT, Hour1 TEXT, Hour2 TEXT, Price_Elasticity FLOAT);'
        cur.execute(sql)
        elasticity = list()
        if test:
            sheetwrite = wbwrite.get_sheet_by_name('Sheet1')
        amount_of_days = length_period/24
        for season in range (0,4):
            print 'season: ', season
            print 'season: ', season
            for day in range(0,amount_of_days):
                if day == startday_weekend or day == startday_weekend+1:
                    print 'in if, and index = ', season*2+1
                    sh=book.sheet_by_index(season*2+1)
                else:
                    print 'in else, and index = ', season*2
                    sh=book.sheet_by_index(season*2)
                for row in range(3,sh.nrows):
                    hour1 = int(sh.cell_value(row, 0)) + 24*day
                    for col in range(1,sh.ncols):
                        if col < 13:
                            if row > 14 + col:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day+1)
                            else:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day)
                            if hour2 > length_period:
                                hour2 = hour2 - length_period
                        else:
                            if col > 11 + row-2:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day-1)
                            else:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day)
                            if hour2 < 1:
                                hour2 = hour2 + length_period
                        value = sh.cell_value(row,col)
                        elasticity.append((season+1,hour1,hour2,value))
                        if test:
                            writecell = sheetwrite.cell(row = hour1, column = hour2)
                            # print value
                            writecell.value = value
        cur.executemany('INSERT INTO Elasticity VALUES (?,?,?,?)', elasticity)
        conn.commit()
        print 'Done elasticities'

    if correction_matrix:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel/auxiliary/correction.xlsx"))
        #sheet 1 = "diagonal"
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS C_Diagonal;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS C_Diagonal (Hour1 TEXT, Hour2 TEXT, Include_Value FLOAT);'
        cur.execute(sql)
        if test:
            sheetwrite = wbwrite.get_sheet_by_name('Sheet2')
        include_value = list()
        amount_of_days = length_period/24
        for day in range(0,amount_of_days):
            for row in range(1,sh.nrows):
                hour1 = int(sh.cell_value(row, 0)) + 24*day
                for col in range(1,sh.ncols):
                    if col < 13:
                        if row > 12 + col:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day+1)
                        else:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day)
                        if hour2 > length_period:
                            hour2 = hour2 - length_period
                    else:
                        if col > 11 + row:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day-1)
                            if hour2 < 1:
                                hour2 = hour2 + length_period
                        else:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day)
                    value = sh.cell_value(row,col)
                    include_value.append((hour1,hour2,value))
                    if test:
                        writecell = sheetwrite.cell(row = hour1, column = hour2)
                        # print value
                        writecell.value = value
        cur.executemany('INSERT INTO C_Diagonal VALUES (?,?,?)', include_value)
        conn.commit()
        print 'Done diagonal'

        #sheet 2 = "LowerT"
        sh = book.sheet_by_index(1)
        sql = 'DROP TABLE IF EXISTS C_LowerT;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS C_LowerT (Hour1 TEXT, Hour2 TEXT, Include_Value FLOAT);'
        cur.execute(sql)
        if test:
            sheetwrite = wbwrite.get_sheet_by_name('Sheet3')
        include_value = list()
        amount_of_days = length_period/24
        for day in range(0,amount_of_days):
            for row in range(1,sh.nrows):
                hour1 = int(sh.cell_value(row, 0)) + 24*day
                for col in range(1,sh.ncols):
                    if col < 13:
                        if row > 12 + col:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day+1)
                        else:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day)
                        if hour2 > length_period:
                            hour2 = hour2 - length_period
                    else:
                        if col > 11 + row:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day-1)
                            if hour2 < 1:
                                hour2 = hour2 + length_period
                        else:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day)
                    value = sh.cell_value(row,col)
                    include_value.append((hour1,hour2,value))
                    if test:
                        writecell = sheetwrite.cell(row = hour1, column = hour2)
                        # print value
                        writecell.value = value
        cur.executemany('INSERT INTO C_LowerT VALUES (?,?,?)', include_value)
        conn.commit()
        print 'Done LowerT'

        #sheet 3 = "UpperT"
        sh = book.sheet_by_index(2)
        sql = 'DROP TABLE IF EXISTS C_UpperT;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS C_UpperT (Hour1 TEXT, Hour2 TEXT, Include_Value FLOAT);'
        cur.execute(sql)
        if test:
            sheetwrite = wbwrite.get_sheet_by_name('Sheet4')
        include_value = list()
        amount_of_days = length_period/24
        for day in range(0,amount_of_days):
            for row in range(1,sh.nrows):
                hour1 = int(sh.cell_value(row, 0)) + 24*day
                for col in range(1,sh.ncols):
                    if col < 13:
                        if row > 12 + col:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day+1)
                        else:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day)
                        if hour2 > length_period:
                            hour2 = hour2 - length_period
                    else:
                        if col > 11 + row:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day-1)
                            if hour2 < 1:
                                hour2 = hour2 + length_period
                        else:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day)
                    value = sh.cell_value(row,col)
                    include_value.append((hour1,hour2,value))
                    if test:
                        writecell = sheetwrite.cell(row = hour1, column = hour2)
                        # print value
                        writecell.value = value
        cur.executemany('INSERT INTO C_UpperT VALUES (?,?,?)', include_value)
        conn.commit()
        print 'Done UpperT'

        #sheet 4 = "linear"
        sh = book.sheet_by_index(3)
        sql = 'DROP TABLE IF EXISTS Linear;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Linear (Hour1 TEXT, Hour2 TEXT, Include_Value FLOAT);'
        cur.execute(sql)
        if test:
            sheetwrite = wbwrite.get_sheet_by_name('Sheet5')
        include_value = list()
        amount_of_days = length_period/24
        for day in range(0,amount_of_days):
            for row in range(1,sh.nrows):
                hour1 = int(sh.cell_value(row, 0)) + 24*day
                for col in range(1,sh.ncols):
                    if col < 13:
                        if row > 12 + col:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day+1)
                        else:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day)
                        if hour2 > length_period:
                            hour2 = hour2 - length_period
                    else:
                        if col > 11 + row:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day-1)
                            if hour2 < 1:
                                hour2 = hour2 + length_period
                        else:
                            hour2 = int(sh.cell_value(0, col)) + 24*(day)
                    value = sh.cell_value(row,col)
                    include_value.append((hour1,hour2,value))
                    if test:
                        writecell = sheetwrite.cell(row = hour1, column = hour2)
                        # print value
                        writecell.value = value
        cur.executemany('INSERT INTO Linear VALUES (?,?,?)', include_value)
        conn.commit()
        print 'Done Linear'


        if test:
            print 'testmatrices written in excel'
            wbwrite.save('excel/TestElasticity.xlsx')

        print 'Done correction matrices'

    if shiftingconstraints:
        minframe = list()
        maxframe = list()
        #frame is amount in two directions from middlepunt that is handled
        frame = 10
        overlapbefore = 1
        overlapafter = 1
        amount_of_overlap_included = 0.5
        sql = 'DROP TABLE IF EXISTS Minframe;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Minframe (Hour1 TEXT, Hour2 TEXT, Inner_frame FLOAT);'
        cur.execute(sql)
        for i in range(1,length_period+1):
            for k in range(1-frame+(i-1),frame+1+i):
                if k<1:
                    newk = length_period+k
                elif k>length_period:
                    newk = k-length_period
                else:
                    newk = k
                minframe.append((i,newk,1))
        sql = 'DROP TABLE IF EXISTS Maxframe;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Maxframe (Hour1 TEXT, Hour2 TEXT, Outer_frame FLOAT);'
        cur.execute(sql)
        for i in range(1,length_period+1):
            for k in range(1-frame-overlapbefore+(i-1),frame+1+overlapafter+i):
                if k<1:
                    newk = length_period+k
                elif k>length_period:
                    newk = k-length_period
                else:
                    newk = k
                if k == frame+overlapafter+i or k == 1-frame-overlapbefore+(i-1):
                    maxframe.append((i,newk,amount_of_overlap_included))
                else:
                    maxframe.append((i,newk,1))
        print minframe
        print maxframe
        cur.executemany('INSERT INTO Minframe VALUES (?,?,?)', minframe)
        cur.executemany('INSERT INTO Maxframe VALUES (?,?,?)', maxframe)
        conn.commit()
        print 'Done shiftingconstraints'


    if demand_ref:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Demand.xlsx"))
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS Demand_energy;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Demand_energy (Year TEXT, Zone TEXT, Demand FLOAT);'
        cur.execute(sql)
        energy = list()
        years = [int(year[0]) for year in cur.execute('SELECT Year from Years').fetchall()]
        # print years
        zones = [zone[0] for zone in cur.execute('SELECT Code from Zone').fetchall()]
        # print zones
        for row in range(3, sh.nrows):
            year = int(sh.cell_value(row, 0))
            if year in years:
                for col in range(1, sh.ncols):
                    zone = sh.cell_value(2, col)
                    if zone in zones:
                        demand = sh.cell_value(row, col)
                        # print year, zone, demand
                        energy.append((year, zone, demand))
        cur.executemany('INSERT INTO Demand_energy VALUES (?,?,?)', energy)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(1)
        sql = 'DROP TABLE IF EXISTS Demand_profile;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Demand_profile (Zone TEXT, Time TEXT, Demand FLOAT);'
        cur.execute(sql)
        profile = list()
        zones = [zone[0] for zone in cur.execute('SELECT Code from Zone').fetchall()]
        #print zones
        for col in range(1, sh.ncols):
            zone = sh.cell_value(2, col)
            for row in range(1, sh.nrows - 2):
                time = '%d' % row
                if zone in zones:
                    demand = sh.cell_value(row + 2, col)
                    # print hour, zone, demand
                    profile.append((zone, time, demand))
        cur.executemany('INSERT INTO Demand_profile VALUES (?,?,?)', profile)
        conn.commit()
        #demand = [zone[0] for zone in cur.execute('SELECT Demand from Demand_profile').fetchall()]
        #plt.plot(demand)
        #plt.show()
        ############################################

        sh = book.sheet_by_index(2)
        sql = 'DROP TABLE IF EXISTS Demand_non_residential;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Demand_non_residential (Zone TEXT, Time TEXT, Demand FLOAT);'
        cur.execute(sql)
        profile = list()
        #print zones
        zone = 'BEL_Z'
        col = 3
        for row in range(1, sh.nrows - 2):
            time = '%d' % row
            demand = sh.cell_value(row-1, col)
            # print hour, zone, demand
            profile.append((zone, time, demand))
        cur.executemany('INSERT INTO Demand_non_residential VALUES (?,?,?)', profile)
        conn.commit()

        #####
        sql = 'DROP TABLE IF EXISTS Demand_residential;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Demand_residential (Zone TEXT, Time TEXT, Demand FLOAT);'
        cur.execute(sql)
        profile = list()
        #print zones
        zone = 'BEL_Z'
        col = 2
        for row in range(1, sh.nrows - 2):
            time = '%d' % row
            demand = sh.cell_value(row-1, col)
            # print hour, zone, demand
            profile.append((zone, time, demand))
        cur.executemany('INSERT INTO Demand_residential VALUES (?,?,?)', profile)
        conn.commit()

        ############################################

        print "Done demand"

    if intermittent_renewables:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Intermittent_renewables.xlsx"))
        sql = 'DROP TABLE IF EXISTS Intermittent_renewables;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Intermittent_renewables (Technology TEXT, Zone TEXT, Time TEXT, Production FLOAT);'
        cur.execute(sql)
        for s in range(book.nsheets):
            sh = book.sheet_by_index(s)
            profile = list()
            technologies = [tech[0] for tech in cur.execute('SELECT Code from Generation_technologies').fetchall()]
            # print technologies
            zones = [zone[0] for zone in cur.execute('SELECT Code from Zone').fetchall()]
            # print zones
            tech = sh.cell_value(0,0)
            if tech in technologies:
                for col in range(1, sh.ncols):
                    zone = sh.cell_value(2, col)
                    for row in range(1, sh.nrows - 2):
                        time = '%d' % row
                        if zone in zones:
                            production = sh.cell_value(row + 2, col)
                            # print tech, hour, zone, production
                            profile.append((tech, zone, time, production))
            cur.executemany('INSERT INTO Intermittent_renewables VALUES (?,?,?,?)', profile)
            conn.commit()
        ############################################
        print "Done intermittent_renewables"

    if reliable_intermittent:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Reliable_intermittent.xlsx"))
        sql = 'DROP TABLE IF EXISTS Reliable_intermittent;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Reliable_intermittent (Technology TEXT, Time TEXT, Zone TEXT, Production FLOAT);'
        cur.execute(sql)
        for s in range(book.nsheets):
            sh = book.sheet_by_index(s)
            profile = list()
            technologies = [tech[0] for tech in cur.execute('SELECT Code from Generation_technologies').fetchall()]
            # print technologies
            zones = [zone[0] for zone in cur.execute('SELECT Code from Zone').fetchall()]
            # print zones
            tech = sh.cell_value(0,0)
            if tech in technologies:
                for row in range(1, sh.nrows - 2):
                    time = '%d' % row
                    for col in range(1, sh.ncols):
                        zone = sh.cell_value(2, col)
                        if zone in zones:
                            production = sh.cell_value(row + 2, col)
                            # print tech, hour, zone, production
                            profile.append((tech, time, zone, production))
            cur.executemany('INSERT INTO Reliable_intermittent VALUES (?,?,?,?)', profile)
            conn.commit()
        ############################################
        print "Done reliable_intermittent"

    if policy:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Policy.xlsx"))
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS Policy_instruments;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Policy_instruments (Include BOOLEAN, Instrument TEXT, Description TEXT);'
        cur.execute(sql)
        instruments = list()
        for rx in range(sh.nrows - 3):
            a = sh.cell_value(rowx=rx + 3, colx=0)
            b = sh.cell_value(rowx=rx + 3, colx=1)
            c = sh.cell_value(rowx=rx + 3, colx=2)
            # print a,b,c
            instruments.append((a, b, c))
        # print areas
        cur.executemany('INSERT INTO Policy_instruments VALUES (?,?,?)', instruments)
        conn.commit()
        ############################################

        sql = 'DROP TABLE IF EXISTS Policy_targets;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Policy_targets (Instrument TEXT, Year TEXT, Target Value);'
        cur.execute(sql)
        years = [year[0] for year in cur.execute('SELECT Year from Years').fetchall()]
        print years
        for s in range(1,book.nsheets):
            sh = book.sheet_by_index(s)
            targets = list()
            instrument = sh.cell_value(0,0)
            instruments = [instr[0] for instr in cur.execute('SELECT Instrument from Policy_instruments').fetchall()]
            # print instruments
            if instrument in instruments:
                for row in range(3, sh.nrows):
                    year = str(int(sh.cell_value(row, 0)))
                    # print year
                    if year in years:
                        target = sh.cell_value(row, 1)
                        # print instrument, year, target
                        targets.append((instrument, year, target))
            cur.executemany('INSERT INTO Policy_targets VALUES (?,?,?)', targets)
            conn.commit()
        ############################################
        print "Done policy"

    if installed_capacities:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Capacity.xlsx"))
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS Capacities;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Capacities (Zone TEXT, Technology TEXT, Year TEXT, Capacity VALUE);'
        cur.execute(sql)
        capacities = list()
        zones = [zone[0] for zone in cur.execute('SELECT Code from Zone').fetchall()]
        #print zones
        technologies = [tech[0] for tech in cur.execute('SELECT Code from Generation_technologies').fetchall()]
        #print technologies
        years = [years[0] for years in cur.execute('SELECT Year from Years').fetchall()]
        #print years
        for cx in range(sh.ncols-2):
            year = str(int(sh.cell_value(2, colx=cx + 2)))
            #print year
            if year in years:
                for rx in range(sh.nrows-3):
                    zone = sh.cell_value(rowx=rx + 3, colx=0)
                    technology = sh.cell_value(rowx=rx + 3, colx=1)
                    if (zone in zones) and (technology in technologies):
                        capacity = sh.cell_value(rowx=rx + 3, colx=cx + 2)
                        # print zone, technology, reserve_type, requirement
                        capacities.append((zone, technology, year, capacity))
        # print probabilistic_reserves
        cur.executemany('INSERT INTO Capacities VALUES (?,?,?,?)', capacities)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(1)
        sql = 'DROP TABLE IF EXISTS Allowed_Generation;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Allowed_Generation (Zone TEXT, Technology TEXT, Allowed BOOLEAN);'
        cur.execute(sql)
        allowed_gen = list()
        technologies = [tech[0] for tech in cur.execute('SELECT Code from Generation_technologies').fetchall()]
        #print technologies
        zones = [zone[0] for zone in cur.execute('SELECT Code from Zone').fetchall()]
        #print zones
        for row in range(3, sh.nrows):
            tech = sh.cell_value(row, 0)
            if tech in technologies:
                for col in range(1, sh.ncols):
                    zone = sh.cell_value(2, col)
                    if zone in zones:
                        allowed = sh.cell_value(row, col)
                        # print tech, zone, allowed
                        allowed_gen.append((zone, tech, allowed))
        cur.executemany('INSERT INTO Allowed_Generation VALUES (?,?,?)', allowed_gen)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(2)
        sql = 'DROP TABLE IF EXISTS Allowed_Storage;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Allowed_Storage (Zone TEXT, Technology TEXT, Allowed BOOLEAN);'
        cur.execute(sql)
        allowed_stor = list()
        technologies = [tech[0] for tech in cur.execute('SELECT Code from Storage_technologies').fetchall()]
        #print technologies
        zones = [zone[0] for zone in cur.execute('SELECT Code from Zone').fetchall()]
        #print zones
        for row in range(3, sh.nrows):
            tech = sh.cell_value(row, 0)
            if tech in technologies:
                for col in range(1, sh.ncols):
                    zone = sh.cell_value(2, col)
                    if zone in zones:
                        allowed = sh.cell_value(row, col)
                        # print tech, zone, allowed
                        allowed_stor.append((zone, tech, allowed))
        cur.executemany('INSERT INTO Allowed_Storage VALUES (?,?,?)', allowed_stor)
        conn.commit()
        ############################################

        print "Done installed capacities"

    if reserves_techn:
        book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Reserves.xlsx"))
        sh = book.sheet_by_index(0)
        sql = 'DROP TABLE IF EXISTS Reserves;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Reserves (Include BOOLEAN, RU BOOLEAN, RUF BOOLEAN, RUA BOOLEAN, RDA BOOLEAN, Type TEXT, Description TEXT);'
        cur.execute(sql)
        reserve_types = list()
        for rx in range(sh.nrows - 3):
            include     = sh.cell_value(rowx=rx + 3, colx=0)
            ru          = sh.cell_value(rowx=rx + 3, colx=1)
            ruf         = sh.cell_value(rowx=rx + 3, colx=2)
            rua         = sh.cell_value(rowx=rx + 3, colx=3)
            rda         = sh.cell_value(rowx=rx + 3, colx=4)
            reserve_type= sh.cell_value(rowx=rx + 3, colx=5)
            description = sh.cell_value(rowx=rx + 3, colx=6)
            # print include, reserve_type, description
            reserve_types.append((include, ru, ruf, rua, rda, reserve_type, description))
        # print areas
        cur.executemany('INSERT INTO Reserves VALUES (?,?,?,?,?,?,?)', reserve_types)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(1)
        sql = 'DROP TABLE IF EXISTS Reserves_deterministic;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Reserves_deterministic (Country Text, Type TEXT, Requirement VALUE);'
        cur.execute(sql)
        deterministic_reserves = list()
        countries = [country[0] for country in cur.execute('SELECT Code from Country').fetchall()]
        # print countries
        reserve_types = [reserves[0] for reserves in cur.execute('SELECT Type from Reserves').fetchall()]
        # print reserve_types
        for cx in range(sh.ncols-1):
            country = sh.cell_value(2, colx=cx + 1)
            if country in countries:
                for rx in range(sh.nrows - 3):
                    reserve_type = sh.cell_value(rowx=rx + 3, colx=0)
                    if reserve_type in reserve_types:
                        requirement = sh.cell_value(rowx=rx + 3, colx=cx + 1)
                        # print country, reserve_type, requirement
                        deterministic_reserves.append((country, reserve_type, requirement))
        # print deterministic_reserves
        cur.executemany('INSERT INTO Reserves_deterministic VALUES (?,?,?)', deterministic_reserves)
        conn.commit()
        ############################################

        sh = book.sheet_by_index(2)
        sql = 'DROP TABLE IF EXISTS Reserves_probabilistic;'
        cur.execute(sql)
        sql = 'CREATE TABLE IF NOT EXISTS Reserves_probabilistic (Country TEXT, Technology TEXT, Type TEXT, Requirement VALUE);'
        cur.execute(sql)
        probabilistic_reserves = list()
        countries = [country[0] for country in cur.execute('SELECT Code from Country').fetchall()]
        #print countries
        technologies = [tech[0] for tech in cur.execute('SELECT Code from Generation_technologies').fetchall()]
        #print technologies
        reserve_types = [reserves[0] for reserves in cur.execute('SELECT Type from Reserves').fetchall()]
        #print reserve_types
        for cx in range(sh.ncols-2):
            technology = sh.cell_value(2, colx=cx + 2)
            if technology in technologies:
                for rx in range(sh.nrows-3):
                    country = sh.cell_value(rowx=rx + 3, colx=0)
                    reserve_type = sh.cell_value(rowx=rx + 3, colx=1)
                    if (country in countries) and (reserve_type in reserve_types):
                        requirement = sh.cell_value(rowx=rx + 3, colx=cx + 2)
                        # print country, technology, reserve_type, requirement
                        probabilistic_reserves.append((country, technology, reserve_type, requirement))
        # print probabilistic_reserves
        cur.executemany('INSERT INTO Reserves_probabilistic VALUES (?,?,?,?)', probabilistic_reserves)
        conn.commit()
        ############################################
        print "Done reserves"


    print "Done"
    conn.close()
